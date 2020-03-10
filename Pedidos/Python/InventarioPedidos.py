# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
import datetime
from openpyxl.utils import get_column_letter

idpedidos = pd.read_excel(r"Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx")

df_pedidos= pd.DataFrame(idpedidos, columns= ['IdPedido'])
#n=len(df_pedidos)

l=df_pedidos.iloc[:, 0].tolist()

#Conexion con la base de fatos
try:
	server = '35.196.201.168' 
	database = 'PETROCHEMICAL_PRUEBAS' 
	username = 'KAIZEN' 
	password = 'SYSERP2016-9#' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto a la BD")
except Exception as e:
	print ("No Conecto a la BD")
	raise e

data = pd.read_sql('Select P.ID_PEDIDO as IdPedido, P.ID_PRODUCTO as IdProducto, P.CANTIDAD as Cantidad, P.CODIGO_PRODUCTO as CodProducto, E.NOMBRE As Estatus_Pedido From PRODUCTO_PEDIDO_MOVIL as P, PEDIDO_MOVIL AS M, ESTATUS_PEDIDO AS E Where P.ID_PEDIDO = M.ID AND M.ID_ESTATUS = E.ID AND  M.ID_ESTATUS=1033  AND P.FACTURABLE = 0 AND P.ID_PEDIDO in (' + ','.join((str(n) for n in l)) + ')',cnxn)
data.to_excel('Inventario.xlsx')

producidos=pd.read_excel(r"Inventario.xlsx")
df_producidos=  pd.DataFrame(producidos, columns= ['IdPedido', 'IdProducto', 'Cantidad', 'CodProducto', 'Estatus_Pedido'])

#################################################################################
#Escribir el Usuario y contrasena
wb = openpyxl.Workbook() 
sheet = wb.active
sheet.cell(row=1, column=1).value ="Idioma"
sheet.cell(row=1, column=2).value ="Usuario"
sheet.cell(row=1, column=3).value ="Contrasena"

count=0
count5=0

lg=(len(df_producidos))+1
for rows in sheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=lg):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="0"
			else:
				row.value="1"	

for rows in sheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=lg):
	for row in rows:
		row.value="automated.test@grupokaizen.net"

for rows in sheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=lg):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb.save("TestDataUser.xlsx") 

usuarios = pd.read_excel(r"TestDataUser.xlsx")
df_usuarios =  pd.DataFrame(usuarios, columns= ['Idioma', 'Usuario', 'Contrasena'])


#################################################################################

df_inventario=pd.concat([df_usuarios, df_producidos], axis=1)

export_excel = df_inventario.to_excel (r'Pedidos/DataExcel/Inventario.xlsx', index = None, header=True)

os.remove('TestDataUser.xlsx')
os.remove('Inventario.xlsx')










