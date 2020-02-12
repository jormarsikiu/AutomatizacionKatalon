# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
import datetime
from openpyxl.utils import get_column_letter

idpedidos = pd.read_excel(r"Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx")

df_pedidos= pd.DataFrame(idpedidos, columns= ['IdPedido', 'CantProductos'])

#n=len(df_pedidos)

l=df_pedidos.iloc[:, 0].tolist()

#Conexion con la base de fatos
try:
	server = '35.196.201.168' 
	database = 'PETROCHEMICAL_PRUEBAS' 
	username = 'KAIZEN' 
	password = 'SYSERP2016-9#' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto a la BD")
except Exception as e:
	print ("No Conecto a la BD")
	raise e

data = pd.read_sql('select P.ID AS IdPedido, P.ID_CLIENTE AS Id_Cliente, E.NOMBRE As Estatus_Pedido from PEDIDO_MOVIL AS P, ESTATUS_PEDIDO AS E WHERE P.ID_ESTATUS = E.ID AND P.ID_ESTATUS=1033 AND P.ID in (' + ','.join((str(n) for n in l)) + ')',cnxn)

data.to_excel('Despachar.xlsx')

pedidoconfirmados=pd.read_excel(r"Despachar.xlsx")

#Generacion de cargos logisticos aleatorios
df_confirmados=  pd.DataFrame(pedidoconfirmados, columns= ['IdPedido', 'Id_Cliente', 'Estatus_Pedido'])

export_excel = df_confirmados.to_excel (r'Despachar.xlsx', index = None, header=True)

#################################################################################
#Escribir el Usuario y contrasena
wb = openpyxl.Workbook() 
sheet = wb.active
sheet.cell(row=1, column=1).value ="Idioma"
sheet.cell(row=1, column=2).value ="Usuario"
sheet.cell(row=1, column=3).value ="Contrasena"

count=0
count5=0

lg=(len(df_confirmados))+1
for rows in sheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=lg):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="0"
			else:
				row.value="1"	

for rows in sheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=lg):
	for row in rows:
		row.value="automated.test@grupokaizen.net"

for rows in sheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=lg):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb.save("TestDataUser.xlsx") 

usuarios = pd.read_excel(r"TestDataUser.xlsx")
df_usuarios =  pd.DataFrame(usuarios, columns= ['Idioma', 'Usuario', 'Contrasena'])


#################################################################################

df_fusion_orden=pd.concat([df_usuarios, df_confirmados, ], axis=1)

df_orden_prod=pd.merge(df_fusion_orden, df_pedidos, on="IdPedido", how="left")

export_excel = df_orden_prod.to_excel(r'Pedidos/DataExcel/Despacho_20pies.xlsx', index = None, header=True)

os.remove('TestDataUser.xlsx')
os.remove('Despachar.xlsx')
