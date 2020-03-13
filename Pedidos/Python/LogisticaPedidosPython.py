# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
import datetime
from openpyxl.utils import get_column_letter
from ConexionBD import *
print(SERVER)

idpedidos = pd.read_excel(r"Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx")

df_pedidos= pd.DataFrame(idpedidos, columns= ['IdPedido'])
#n=len(df_pedidos)

l=df_pedidos.iloc[:, 0].tolist()

#Conexion con la base de fatos
try:
	server = SERVER
	database = DATABASE
	username = USERNAME
	password = PASSWORD
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto a la BD")
except Exception as e:
	print ("No Conecto a la BD")
	raise e

data = pd.read_sql('select P.ID AS IdPedido, P.ID_CLIENTE AS Id_Cliente, E.NOMBRE As Estatus_Pedido from PEDIDO_MOVIL AS P, ESTATUS_PEDIDO AS E WHERE P.ID_ESTATUS = E.ID AND P.ID_ESTATUS=2 AND P.ID in (' + ','.join((str(n) for n in l)) + ')',cnxn)
data1 = pd.read_sql('select ID as IDCargo, NOMBRE as NombreCargo from producto where ID_GRUPO=1',cnxn)
data2 = pd.read_sql('select ID as IDCargo2, NOMBRE as NombreCargo2 from producto where ID_GRUPO=1',cnxn)

data.to_excel('TestCargosInv.xlsx')
data1.to_excel('TestDataCargoLogistico.xlsx')
data2.to_excel('TestDataCargoLogistico2.xlsx')

pedidoaprobados=pd.read_excel(r"TestCargosInv.xlsx")
cargologistico = pd.read_excel(r"TestDataCargoLogistico.xlsx")
cargologistico2 = pd.read_excel(r"TestDataCargoLogistico2.xlsx")

#Generacon de cargos logisticos aleatorios
df_pedidoaprobados=  pd.DataFrame(pedidoaprobados, columns= ['IdPedido', 'Id_Cliente', 'Estatus_Pedido'])
df_cargo = pd.DataFrame(cargologistico, columns= ['IDCargo', 'NombreCargo'])
df_cargo2 = pd.DataFrame(cargologistico2, columns= ['IDCargo2', 'NombreCargo2'])

nc=len(df_cargo)
df_aleatorio_cargo = df_cargo.sample(nc)
df_aleatorio_cargo2 = df_cargo2.sample(nc)
df_aleatorio_cargo = df_aleatorio_cargo.reset_index()
df_aleatorio_cargo2 = df_aleatorio_cargo2.reset_index()

export_excel = df_aleatorio_cargo.to_excel (r'TestDataCargoLogistico.xlsx', index = None, header=True)
export_excel2 = df_aleatorio_cargo2.to_excel (r'TestDataCargoLogistico2.xlsx', index = None, header=True)

df_fusion=pd.concat([df_aleatorio_cargo, df_aleatorio_cargo2], axis=1)

#################################################################################
#Determina que la cantidad de registros de clientes sea igual a la cantidad de productos
#para que no se desborde la tabla
long1=len(df_pedidoaprobados)
long2=len(df_aleatorio_cargo)

if (long1<=long2):
	n=long1
else:
	n=long2

df_fusion = df_fusion.iloc[0:n]
export_excel4 = df_fusion.to_excel(r'CargoLogistico.xlsx', header=True)

#################################################################################
#Escribir el Usuario y contrasena
wb = openpyxl.Workbook() 
sheet = wb.active
sheet.cell(row=1, column=1).value ="Idioma"
sheet.cell(row=1, column=2).value ="Usuario"
sheet.cell(row=1, column=3).value ="Contrasena"

count=0
count5=0

lg=(len(df_fusion))+1
for rows in sheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=lg):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="0"
			else:
				row.value="1"	

for rows in sheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=lg):
	for row in rows:
		row.value="automated.test@grupokaizen.net"

for rows in sheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=lg):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb.save("TestDataUser.xlsx") 

usuarios = pd.read_excel(r"TestDataUser.xlsx")
df_usuarios =  pd.DataFrame(usuarios, columns= ['Idioma', 'Usuario', 'Contrasena'])
#################################################################################
#Generar la union de los dos archivos de los cargos logisticos aleatorios 

cargologistico3 = pd.read_excel(r"CargoLogistico.xlsx")
df_logistico=pd.DataFrame(cargologistico3, columns= ['IDCargo', 'NombreCargo', 'IDCargo2', 'NombreCargo2'], index = None)
df_pedidoaprobados = df_pedidoaprobados.reset_index()

df_fusion_logistico=pd.concat([df_usuarios, df_pedidoaprobados, df_logistico], axis=1)
df_fusion_logistico = df_fusion_logistico.drop(['index'], axis=1)
export_excel5 = df_fusion_logistico.to_excel(r'Pedidos/DataExcel/Cargos_Logisticos.xlsx', index = None, header=True)


os.remove('TestCargosInv.xlsx')
os.remove('TestDataCargoLogistico.xlsx')
os.remove('TestDataCargoLogistico2.xlsx')
os.remove('CargoLogistico.xlsx')
os.remove('TestDataUser.xlsx')