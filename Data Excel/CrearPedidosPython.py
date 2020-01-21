# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
from openpyxl.utils import get_column_letter
from datetime import datetime


#Conexion con la base de fatos
try:
	server = '35.196.201.168' 
	database = 'PETROCHEMICAL_PRUEBAS' 
	username = 'KAIZEN' 
	password = 'SYSERP2016-9#' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto")
except Exception as e:
	print ("No Conecto")
	raise e

#Consultas SQL a la base de datos
#data = pd.read_sql('Select C.ID AS Nombre, D.ID AS Direccion, G.CODIGO AS Contenedor from CLIENTE as C, DIRECCION_CLIENTE as D, CONTENEDOR as G Where C.ID = D.ID_CLIENTE',cnxn)
data = pd.read_sql('Select C.ID  AS ID_Cliente, C.NOMBRE, D.ID AS ID_Direccion, D.Nombre from CLIENTE as C, DIRECCION_CLIENTE as D, CONTENEDOR as G Where C.ID = D.ID_CLIENTE AND C.ELIMINADO=0 AND C.ACTIVO=1 AND D.ACTIVO=1 AND D.ELIMINADO=0',cnxn)
data2 = pd.read_sql('Select  O.ID AS OpCompra, P.ID AS Presentacion, T.ID AS Producto, S.CANTIDAD AS Cantidad from OPCION_COMPRA as O, OPCION_COMPRA_PRESENTACION as S,PRESENTACION as P, PRODUCTO as T, PRESENTACION_PRODUCTO as A  where O.ID = S.ID_OPCION AND P.ID = S.ID_PRESENTACION AND A.ID_PRODUCTO=T.ID AND A.ID_PRESENTACION=S.ID_PRESENTACION AND O.ELIMINADO=0 AND O.ACTIVO=1 AND P.ELIMINADO=0 AND P.ACTIVO=1 AND T.ELIMINADO=0 AND T.ACTIVO=1 AND T.EN_VENTA=1',cnxn)
data3 = pd.read_sql('Select  O.ID AS OpCompra, P.ID AS Presentacion2, T.ID AS Producto2, S.CANTIDAD AS Cantidad2 from OPCION_COMPRA as O, OPCION_COMPRA_PRESENTACION as S,PRESENTACION as P, PRODUCTO as T, PRESENTACION_PRODUCTO as A  where O.ID = S.ID_OPCION AND P.ID = S.ID_PRESENTACION AND A.ID_PRODUCTO=T.ID AND A.ID_PRESENTACION=S.ID_PRESENTACION AND O.ELIMINADO=0 AND O.ACTIVO=1 AND P.ELIMINADO=0 AND P.ACTIVO=1 AND T.ELIMINADO=0 AND T.ACTIVO=1 AND T.EN_VENTA=1',cnxn)


#Se guardan las consultas en el excel 
data.to_excel('TestDataClient.xlsx')
data2.to_excel('TestDataProduct.xlsx')
data3.to_excel('TestDataProduct2.xlsx')


############################################
#Leer el tamaño del archivo
wb = openpyxl.load_workbook('TestDataClient.xlsx')
sheet = wb.worksheets[0]
n=len(sheet['A'])
print ("Guardo")

############################################

#Escribir en celdas individuales
#sheet2["A1"]="Incoterms"
#sheet2.cell(row=1, column =1).value ="Incoterms"

############################################
#Escribir el CIF y EXW
wb2 = openpyxl.Workbook() 
sheet2 = wb2.active
sheet2.cell(row=1, column=1).value ="Incoterms"

count=0
for rows in sheet2.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="CIF"
			else:
				row.value="EXW"

wb2.save("TestDataIncoterms.xlsx") 

################################################
#Escribir el Idioma
wb3 = openpyxl.Workbook() 
sheet3 = wb3.active
sheet3.cell(row=1, column=1).value ="Idioma"

count=0
for rows in sheet3.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="0"
			else:
				row.value="1"
wb3.save("TestDataIdioma.xlsx") 

############################################

#Escribir el Contenedor
wb4 = openpyxl.Workbook() 
sheet4 = wb4.active
sheet4.cell(row=1, column=1).value ="Contenedor"

count=0
for rows in sheet4.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="20"
			else:
				row.value="20"
wb4.save("TestDataContenedor.xlsx") 

############################################
#Escribir el Usuario y contrasena
wb5 = openpyxl.Workbook() 
sheet5 = wb5.active
sheet5.cell(row=1, column=1).value ="Usuario"
sheet5.cell(row=1, column=2).value ="Contrasena"
count=0

for rows in sheet5.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
		row.value="automated.test@grupokaizen.net"


for rows in sheet5.iter_cols(min_col=2, max_col=2, min_row=2, max_row=n):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb5.save("TestDataUser.xlsx") 

################################################

#Leemos los excel 
idioma = pd.read_excel(r"TestDataIdioma.xlsx")
usuarios = pd.read_excel(r"TestDataUser.xlsx")
cliente = pd.read_excel(r"TestDataClient.xlsx")
contenedor = pd.read_excel(r"TestDataContenedor.xlsx") 
producto = pd.read_excel(r"TestDataProduct.xlsx")
producto2 = pd.read_excel(r"TestDataProduct2.xlsx")
incoterm = pd.read_excel(r"TestDataIncoterms.xlsx")

#Se crean dataframes a partir del excel
df_idioma= pd.DataFrame(idioma, columns= ['Idioma'])
df_usuarios =  pd.DataFrame(usuarios, columns= ['Usuario', 'Contrasena'])
df_cliente = pd.DataFrame(cliente, columns= ['ID_Cliente', 'ID_Direccion'])
df_icoterms= pd.DataFrame(incoterm, columns= ['Incoterms'])
df_contenedor = pd.DataFrame(contenedor, columns= ['Contenedor'])

#Se Crea un dataframe con las columnas de los dataframes anteriores para productos
df_producto1= pd.DataFrame(producto, columns= ['OpCompra','Presentacion', 'Producto', 'Cantidad'])
df_producto2= pd.DataFrame(producto2, columns= ['OpCompra','Presentacion2', 'Producto2', 'Cantidad2'])

#Se mezclan los productos 1 y 2
df_producto2=pd.merge(df_producto1, df_producto2, on="OpCompra", how="left")
df_producto2['OpCompra2'] = df_producto2['OpCompra']

long1=len(df_usuarios)
long2=len(df_producto2)

#Determina cuales opciones de compra tienen un solo producto y coloca en 0 los valores del segundo producto
for i in range (0, len(df_producto2)):
	if (df_producto2.loc[i, 'OpCompra']<13):
		df_producto2.loc[i, 'Presentacion2']=0
		df_producto2.loc[i, 'Producto2']=0
		df_producto2.loc[i, 'Cantidad2']=0


#Determina que la cantidad de registros de clientes sea igual a la cantidad de productos
#para que no se desborde la tabla
if (long1<=long2):
	n=long1
else:
	n=long2

df_producto2 = df_producto2.iloc[0:n]

#Concatenar los dataframe en uno 
df_fusion=pd.concat([df_idioma, df_usuarios, df_cliente, df_contenedor, df_icoterms, df_producto2], axis=1)
export_excel = df_fusion.to_excel (r'Testfusion.xlsx', index = None, header=True)

#Crear dataframe con filas aleatorias
na=len(df_fusion)
df_aleatorio = df_fusion.sample(na)

#Determinar si se debe iterar 1 o 2 productos
for j in range (0, len(df_aleatorio)):
	if ((df_aleatorio.loc[j, 'Presentacion2']==0 ) &  (df_aleatorio.loc[j, 'Producto2']==0)):
		df_aleatorio.loc[j, 'Iterar'] = 1
	else:
		df_aleatorio.loc[j, 'Iterar'] = 2

#Determina cual prueba es exitosa y cual fallaria
for i in range (0, len(df_aleatorio)):
	if ((df_aleatorio.loc[i,'OpCompra']==df_aleatorio.loc[i,'OpCompra2']) & (df_aleatorio.loc[i,'Presentacion']!=df_aleatorio.loc[i,'Presentacion2'])):
		df_aleatorio.loc[i, 'Resultado'] = "Prueba exitosa"
	else:
		df_aleatorio.loc[i, 'Resultado'] = "Prueba fallida"

export_excel2 = df_aleatorio.to_excel (r'Crear_Pedido.xlsx', index = None, header=True)

#Ajustar columnas
wb6 = openpyxl.load_workbook('Crear_Pedido.xlsx')
sheet6 = wb6.worksheets[0]

sheet6.column_dimensions['B'].width = 20
sheet6.column_dimensions['C'].width = 20
sheet6.column_dimensions['D'].width = 15
sheet6.column_dimensions['E'].width = 15
sheet6.column_dimensions['F'].width = 15
sheet6.column_dimensions['G'].width = 15
sheet6.column_dimensions['H'].width = 15
sheet6.column_dimensions['I'].width = 15
sheet6.column_dimensions['J'].width = 15
sheet6.column_dimensions['K'].width = 15
sheet6.column_dimensions['L'].width = 15
sheet6.column_dimensions['M'].width = 15
sheet6.column_dimensions['N'].width = 15
sheet6.column_dimensions['O'].width = 15
sheet6.column_dimensions['P'].width = 15

wb6.save('Crear_Pedido.xlsx') 

#Borrado de archivos
os.remove('TestDataIdioma.xlsx')
os.remove('TestDataClient.xlsx')
os.remove('TestDataContenedor.xlsx') 
os.remove('TestDataIncoterms.xlsx')
os.remove('TestDataProduct.xlsx')
os.remove('TestDataProduct2.xlsx')
os.remove('Testfusion.xlsx')
os.remove('TestDataUser.xlsx')


