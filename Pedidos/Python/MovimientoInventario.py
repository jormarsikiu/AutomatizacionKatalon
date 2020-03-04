# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
import random
import math
import numpy as np


productos_almacen = pd.read_excel(r"Pedidos/DataExcel/Inventario/ProductoAlmacen.xlsx")
df_productos_almacen= pd.DataFrame(productos_almacen, columns= ['IdProducto'])
l=df_productos_almacen.iloc[:, 0].tolist()

#Conexion con la base de fatos
try:
	server = '35.196.201.168' 
	database = 'PETROCHEMICAL_PRUEBAS' 
	username = 'KAIZEN' 
	password = 'SYSERP2016-9#' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto a la BD")
except Exception as e:
	print ("No Conecto a la BD")
	raise e
#P.ID=PA.ID_PRODUCTO AND PA.ID_ALMACEN=A.ID AND
data = pd.read_sql('SELECT T.NOMBRE AS Concepto, P.VALOR AS Motivo FROM PARAMETRO AS P, TIPO_PARAMETRO AS T WHERE T.ID = P.ID_TIPO_PARAMETRO AND P.ID_TIPO_PARAMETRO IN (17, 18, 19) ORDER BY T.ID',cnxn)
data2 = pd.read_sql('SELECT P.ID as IdProducto, P.CODIGO as CodProducto, P.NOMBRE as Producto, P.EXISTENCIA as Existencia, P.EN_VENTA as EnVenta, P.ID_GRUPO as IdGrupo FROM PRODUCTO AS P WHERE P.ID_GRUPO IN (6, 7, 32) AND P.ID IN (' + ','.join((str(n) for n in l)) + ')',cnxn)
data3=pd.read_sql('SELECT ID_PRODUCTO AS IdProducto, ID_ALMACEN AS IdAlmacen1, CANTIDAD_ACTUAL AS CantidadActual1 FROM PRODUCTO_ALMACEN WHERE ID_ALMACEN IN (1,2,6,7) AND ID_PRODUCTO IN (' + ','.join((str(n) for n in l)) + ')',cnxn)
data4=pd.read_sql('SELECT ID_PRODUCTO AS IdProducto, ID_ALMACEN AS IdAlmacen2, CANTIDAD_ACTUAL AS CantidadActual2 FROM PRODUCTO_ALMACEN WHERE ID_ALMACEN IN (1,2,6,7) AND ID_PRODUCTO IN (' + ','.join((str(n) for n in l)) + ')',cnxn)
data5 = pd.read_sql('SELECT P.ID as IdProducto, P.CODIGO as CodProducto, P.NOMBRE as Producto, P.EXISTENCIA as Existencia, P.EN_VENTA as EnVenta, P.ID_GRUPO as IdGrupo FROM PRODUCTO AS P WHERE P.ID_GRUPO =7 AND P.EN_VENTA =1 AND P.ID IN (' + ','.join((str(n) for n in l)) + ')',cnxn)

data.to_excel('Concepto.xlsx')
data2.to_excel('Productos.xlsx')
data3.to_excel('Almacen1.xlsx')
data4.to_excel('Almacen2.xlsx')
data5.to_excel('ProductosVenta.xlsx')

concepto=pd.read_excel(r"Concepto.xlsx")
df_concepto=  pd.DataFrame(concepto, columns= ['Concepto', 'Motivo'])
n=len(df_concepto)
df_concepto = df_concepto.sample(n)

#Se elimina la fila Venta de Productos para agregarla con un producto en venta
for i in range (0, len(df_concepto)):
	if (df_concepto.loc[i, 'Motivo']=='VentaProducto'):
		df_concepto=df_concepto.drop([i],axis=0)
	elif (df_concepto.loc[i, 'Motivo']=='ParaDespacho'):
		df_concepto=df_concepto.drop([i],axis=0)

#Se limitan los productos a la cantidad de conceptos
producto=pd.read_excel(r"Productos.xlsx")
df_producto=  pd.DataFrame(producto, columns= ['IdProducto', 'IdGrupo', 'CodProducto', 'Producto', 'Existencia', 'EnVenta'])
df_producto = df_producto.iloc[0:9]

#Se extraen los productos en venta para agregarlos al motivo de venta por productos
#Se aleatorian para escoger solo uno
productosventa=pd.read_excel(r"ProductosVenta.xlsx")
df_producto_venta=  pd.DataFrame(productosventa, columns= ['IdProducto', 'IdGrupo', 'CodProducto', 'Producto', 'Existencia', 'EnVenta'])
np=len(df_producto_venta)
df_producto_venta = df_producto_venta.sample(np)
df_producto_venta = df_producto_venta.iloc[0:2]


#Se concatenan los todos los productos con el producto en venta en fusion
df_fusion=pd.concat([df_producto, df_producto_venta], axis=0, ignore_index=True)

#Cambiar el nombre del grupo
for i in range (0, len(df_fusion)):
	if (df_fusion.loc[i, 'IdGrupo']==6):
		df_fusion.loc[i, 'IdGrupo'] = 'Materia Prima'
	elif (df_fusion.loc[i, 'IdGrupo']==7):
		df_fusion.loc[i, 'IdGrupo'] = 'Producto Terminado'
	elif (df_fusion.loc[i, 'IdGrupo']==32):
		df_fusion.loc[i, 'IdGrupo'] = 'MP: Liquidos'

#Se cambia el valor de En Venta por numeros
for i in range (0, len(df_fusion)):
	if (df_fusion.loc[i, 'EnVenta']==True):
		df_fusion.loc[i, 'EnVenta']=1
	elif (df_fusion.loc[i, 'EnVenta']==False):
		df_fusion.loc[i, 'EnVenta']=0

#Se resetean los dataframe de conceptos y fusion para unirlos
df_concepto = df_concepto.reset_index(drop=True)
df_fusion = df_fusion.reset_index(drop=True)
df_mov=pd.concat([df_concepto, df_fusion], axis=1)


#Se agrega el motivo de Venta de Productos en el producto que esta a la venta 

a=(len(df_mov))-3
b=(len(df_mov))-2
c=(len(df_mov))-1
d=len(df_mov)


for i in range (9, 10):
	if (df_mov['Concepto'].isnull().any()):
		df_mov.loc[i, 'Concepto'] = 'SalidaDeMercancias'
		df_mov.loc[i, 'Motivo'] = 'VentaProducto'

for i in range (10, 11):
	if (df_mov['Concepto'].isnull().any()):
		df_mov.loc[i, 'Concepto'] = 'TransferenciaDeMercancias'
		df_mov.loc[i, 'Motivo'] = 'ParaDespacho'

	

#Se seleccionan los productos con los almacenes y la cantidad que hay dentro de ellos
almacen1 = pd.read_excel(r"Almacen1.xlsx")
almacen2 = pd.read_excel(r"Almacen2.xlsx")
df_almacen1= pd.DataFrame(almacen1, columns= ['IdProducto','IdAlmacen1', 'CantidadActual1'])
df_almacen2= pd.DataFrame(almacen2, columns= ['IdProducto','IdAlmacen2', 'CantidadActual2'])

#Se mezclan los almacenes de forma horizontal segun el Id del Producto
df_almacenes=pd.merge(df_almacen1, df_almacen2, on="IdProducto", how="left")

#Se sacan los almacenes que son iguales para el mismo producto que se generan en la mezcla
df_aleatorio2=pd.DataFrame()
for i in range (0, len(df_almacenes)):
	if (df_almacenes.loc[i, 'IdAlmacen1']!=df_almacenes.loc[i, 'IdAlmacen2']):
		df_aleatorio2.loc[i, 'IdProducto']=df_almacenes.loc[i, 'IdProducto']
		df_aleatorio2.loc[i, 'IdAlmacen1']=df_almacenes.loc[i, 'IdAlmacen1']
		df_aleatorio2.loc[i, 'CantidadActual1']=df_almacenes.loc[i, 'CantidadActual1']
		df_aleatorio2.loc[i, 'IdAlmacen2']=df_almacenes.loc[i, 'IdAlmacen2']
		df_aleatorio2.loc[i, 'CantidadActual2']=df_almacenes.loc[i, 'CantidadActual2']

export_excel = df_aleatorio2.to_excel (r'Almacenes.xlsx', index = None, header=True)

#Se mezclan los productos con los almacenes 
df_mov_inventario=pd.merge(df_mov, df_aleatorio2, on="IdProducto", how="left")
export_excel = df_mov_inventario.to_excel (r'Movimiento.xlsx', index = None, header=True)

#################################################################################
#Escribir el Usuario y contrasena
wb = openpyxl.Workbook() 
sheet = wb.active
sheet.cell(row=1, column=1).value ="Idioma"
sheet.cell(row=1, column=2).value ="Usuario"
sheet.cell(row=1, column=3).value ="Contrasena"

count=0
count5=0
#No acepta multidioma aun 
lg=(len(df_mov_inventario))+1
for rows in sheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=lg):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="1" 
			else:
				row.value="1"	

for rows in sheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=lg):
	for row in rows:
		row.value="automated.test@grupokaizen.net"

for rows in sheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=lg):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb.save("TestDataUser.xlsx") 

usuarios = pd.read_excel(r"TestDataUser.xlsx")
df_usuarios =  pd.DataFrame(usuarios, columns= ['Idioma', 'Usuario', 'Contrasena'])


#################################################################################
#Unir los usuarios con los producto
df_mov_inventario=df_mov_inventario.reset_index(drop=True)
df_movinventario2=pd.concat([df_usuarios, df_mov_inventario], axis=1)


#Calcular la cantidad a mover
df_n= pd.DataFrame(df_movinventario2, columns= ['CantidadActual1'])

#Convertir la columna CantidadActual1 en una lista (resultado numeros decimales)
lista=df_n.iloc[:, 0].tolist()

list1 = [0.0 if math.isnan(x) else x for x in lista]

#Convertir la lista en numeros enteros
for i in range(len(list1)):
	list1[i] = int(list1[i])

#Convertir en 0 los numeros negativos para que no se saque mas del almacen 
list_salida = [0 if i < 0 else i for i in list1]
list_tranf = [0 if i < 0 else i for i in list1]

#Convertir los 0 en 1
for i in range(len(list_salida)):
	if (list_salida[i] == 0):
		list_salida[i] = 1

#Para los conceptos de entrada o salida debe ser mayor que 0 y para transferencia puede ser 0 
#Se limita la CantidadMovida con el limite de la CantidadActual1 y se retornan aleatorios
for i in range (0, len(df_movinventario2)):
	if((df_movinventario2.loc[i, 'Concepto']=='TransferenciaDeMercancias')& (df_movinventario2.loc[i, 'CantidadActual1']==0)):
		df_movinventario2.loc[i, 'CantidadAMover'] = 0
	elif((df_movinventario2.loc[i, 'Concepto']=='TransferenciaDeMercancias')& (df_movinventario2.loc[i, 'CantidadActual1']>0)):
		df_movinventario2.loc[i, 'CantidadAMover'] = random.randint(1, list_salida[i])
	else:
		df_movinventario2.loc[i, 'CantidadAMover'] = random.randint(1, list_salida[i])

#Los conceptos que no son de transferencia no tienen almacen 2
for j in range (0, len(df_movinventario2)):
	if (df_movinventario2.loc[j, 'Concepto']!='TransferenciaDeMercancias'):
		df_movinventario2.loc[j, 'IdAlmacen2'] = 'NA'
		df_movinventario2.loc[j, 'CantidadActual2'] = 'NA'

#Se eliminan los registros duplicados
#for j in range (0, len(df_movinventario2)):
	#if (df_movinventario2.loc[j, 'Concepto']!='TransferenciaDeMercancias'):
		#df_movinventario3 = df_movinventario2.drop_duplicates(['IdProducto', 'Existencia'], keep='first')
	#df_movinventario2 = df_movinventario2.drop_duplicates(subset='IdProducto', keep='first')
		#df_movinventario2.drop_duplicates(['IdProducto', 'Existencia'], keep='first')



#print(len(df_movinventario2))
#df_movinventario2 = df_movinventario2.reset_index(drop=True)
#df_movinventario3 = df_movinventario2.drop_duplicates(subset='IdProducto', keep='first')
#nc=len(df_movinventario2)
#df_movinventario3 = df_movinventario2.sample(nc)
#df_movinventario3 = df_movinventario3.reset_index(drop=True)
export_excel = df_movinventario2.to_excel (r'Pedidos/DataExcel/Inventario/MovimientoDeInventario.xlsx', index = None, header=True)

os.remove('Concepto.xlsx')
os.remove('Productos.xlsx')
os.remove('ProductosVenta.xlsx')
os.remove('Almacen1.xlsx')
os.remove('Almacen2.xlsx')
os.remove('Almacenes.xlsx')
os.remove('TestDataUser.xlsx')
os.remove('Movimiento.xlsx')

#Se valida y se crea el archivo que necesita Katalon para guardar los pedidos
if os.path.isfile("Pedidos/DataExcel/Inventario/LogReporteInventarioKatalon.xlsx"):
    print("Ya esta creado: Pedidos/DataExcel/Inventario/LogReporteInventarioKatalon.xlsx")
else:
	wb7 = openpyxl.Workbook() 
	wb7.save('Pedidos/DataExcel/Inventario/LogReporteInventarioKatalon.xlsx') 
	print("Archivo creado: Pedidos/DataExcel/Inventario/LogReporteInventarioKatalon.xlsx")






