# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
from openpyxl.utils import get_column_letter
from datetime import datetime
import random
import numpy as np
import xlsxwriter
from ConexionBD import *
print(SERVER)


#Conexion con la base de fatos
try:
	server = SERVER
	database = DATABASE
	username = USERNAME
	password = PASSWORD
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto a la BD")
except Exception as e:
	print ("No Conecto a la BD")
	raise e

#Consultas SQL a la base de datos
data = pd.read_sql('Select C.ID as IdCliente, D.ID AS IdDireccion, ID_PUERTO AS IdPuerto, CAPACIDAD_DESCARGA AS CapacidadDescarga, (CAPACIDAD_DESCARGA - (SELECT PESO_TARA FROM CONTENEDOR WHERE ID = 4)) AS CapacidadDisponible From CLIENTE as C, DIRECCION_CLIENTE as D, PUERTO_CLIENTE as PC, PUERTO as O WHERE D.ID_CLIENTE = C.ID AND C.ID = PC.ID_CLIENTE AND C.ELIMINADO=0 AND C.ACTIVO=1 AND O.ID=PC.ID_PUERTO and O.ACTIVO=1 AND O.DE_SALIDA=0 AND D.ACTIVO = 1 AND D.ELIMINADO=0',cnxn)
data2 = pd.read_sql('Select PROD.ID AS ProductoSoportaPeso, PROD.PESO AS PesoProducto1 from PRESENTACION as P, PRODUCTO as PROD, PRESENTACION_PRODUCTO as PP where P.ID = 3 AND PROD.ID = PP.ID_PRODUCTO AND P.ID = PP.ID_PRESENTACION AND PROD.ELIMINADO=0 AND PROD.ACTIVO=1 AND P.ELIMINADO=0 AND P.ACTIVO=1 AND PROD.EN_VENTA=1',cnxn)
data3 = pd.read_sql('Select PROD.ID AS ProductoNoSoportaPeso, PROD.PESO AS PesoProducto2 from PRESENTACION as P, PRODUCTO as PROD, PRESENTACION_PRODUCTO as PP where P.ID != 3 AND PROD.ID = PP.ID_PRODUCTO AND  P.ID = PP.ID_PRESENTACION  AND PROD.ELIMINADO=0 AND PROD.ACTIVO=1 AND P.ELIMINADO=0 AND P.ACTIVO=1 AND PROD.EN_VENTA=1',cnxn)

#Se guardan las consultas en el excel 
data.to_excel('TestDataClient.xlsx')
data2.to_excel('TestDataProductSop.xlsx')
data3.to_excel('TestDataProductNo.xlsx')

#Leemos los excel 
cliente = pd.read_excel(r"TestDataClient.xlsx")
productoSoporta = pd.read_excel(r"TestDataProductSop.xlsx")
productoNoSoporta = pd.read_excel(r"TestDataProductNo.xlsx")

#Se crean dataframes a partir del excel
df_cliente = pd.DataFrame(cliente, columns= ['IdCliente', 'IdDireccion', 'IdPuerto', 'CapacidadDescarga', 'CapacidadDisponible'])
df_productoSoporta = pd.DataFrame(productoSoporta, columns= ['ProductoSoportaPeso', 'PesoProducto1'])
df_productoNoSoporta =  pd.DataFrame(productoNoSoporta, columns= ['ProductoNoSoportaPeso','PesoProducto2'])

#Se determina la longitud de filas de dataframe para que no se desborde
long1=len(df_cliente)
long2=len(df_productoSoporta)
long3=len(df_productoNoSoporta)
"""
if(long1 < long2 and long1 < long3):
	n=long1
else:
	if(long2 < long1 and long2 < long3):
		n=long2
	else:
		n=long3
"""

if (long1<=long2):
	n=long1
else:
	n=long2

#Se asignan los tamanos a los dataframes para evitar el desborde
df_cliente = df_cliente.iloc[0:n]
df_productoSoporta = df_productoSoporta.iloc[0:n]
df_productoNoSoporta = df_productoNoSoporta.iloc[0:n]

#Concatenar los dataframe en uno 

df_fusion=pd.concat([df_cliente, df_productoSoporta, df_productoNoSoporta], axis=1)
export_excel = df_fusion.to_excel (r'Testfusion.xlsx', index = None, header=True)
n=n+1

############################################
#Escribir el CIF y EXW
wb2 = openpyxl.Workbook() 
sheet2 = wb2.active
sheet2.cell(row=1, column=1).value ="Incoterms"

count=0
for rows in sheet2.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="CIF"
			else:
				row.value="EXW"

wb2.save("TestDataIncoterms.xlsx") 

################################################
#Escribir el Idioma
wb3 = openpyxl.Workbook() 
sheet3 = wb3.active
sheet3.cell(row=1, column=1).value ="Idioma"

count=0
for rows in sheet3.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="0"
			else:
				row.value="1"
wb3.save("TestDataIdioma.xlsx") 

############################################

#Escribir el Contenedor
wb4 = openpyxl.Workbook() 
sheet4 = wb4.active
sheet4.cell(row=1, column=1).value ="Contenedor"

count=0
for rows in sheet4.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="40"
			else:
				row.value="40"
wb4.save("TestDataContenedor.xlsx") 

############################################
#Escribir el Usuario y contrasena
wb5 = openpyxl.Workbook() 
sheet5 = wb5.active
sheet5.cell(row=1, column=1).value ="Usuario"
sheet5.cell(row=1, column=2).value ="Contrasena"
count=0

for rows in sheet5.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
		row.value="automated.test@grupokaizen.net"


for rows in sheet5.iter_cols(min_col=2, max_col=2, min_row=2, max_row=n):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb5.save("TestDataUser.xlsx") 

###############################################################################
#Leemos los excel 
idioma = pd.read_excel(r"TestDataIdioma.xlsx")
usuarios = pd.read_excel(r"TestDataUser.xlsx")
contenedor = pd.read_excel(r"TestDataContenedor.xlsx") 
incoterm = pd.read_excel(r"TestDataIncoterms.xlsx")

#Se crean dataframes a partir del excel
df_idioma= pd.DataFrame(idioma, columns= ['Idioma'])
df_usuarios =  pd.DataFrame(usuarios, columns= ['Usuario', 'Contrasena'])
df_icoterms= pd.DataFrame(incoterm, columns= ['Incoterms'])
df_contenedor = pd.DataFrame(contenedor, columns= ['Contenedor'])

#Concatenar los dataframe en uno 
df_fusion=pd.concat([df_idioma, df_usuarios, df_contenedor, df_icoterms, df_fusion], axis=1)
export_excel = df_fusion.to_excel (r'Testfusion.xlsx', index = None, header=True)


#Crear dataframe con filas aleatorias
na=len(df_fusion)
df_aleatorio = df_fusion.sample(na)

df_aleatorio['ProductoNoSoportaPeso'] = df_aleatorio['ProductoNoSoportaPeso'].replace(np.nan, 0)
df_aleatorio['PesoProducto2'] = df_aleatorio['PesoProducto2'].replace(np.nan, 0)

#Determinar si se debe iterar 1 o 2 productos
for j in range (0, len(df_aleatorio)):
	if (df_aleatorio.loc[j, 'ProductoNoSoportaPeso']==0):
		df_aleatorio.loc[j, 'Iterar'] = 1
	else:
		df_aleatorio.loc[j, 'Iterar'] = 2

#Determinar se marcan los exitosos
for j in range (0, len(df_aleatorio)):
		df_aleatorio.loc[j, 'Resultado'] = "Prueba exitosa"


#Estadsticas: Crear el borrador al 20% de los pedidos
df_aleatorio2=pd.DataFrame()
df_aleatorio2=df_aleatorio
longitud=1+len(df_aleatorio2)
porcentaje = longitud*(0.20)
porcentaje =int(porcentaje)
longitud2=porcentaje + longitud

#Selecciona data aleatoria para hacer el pedido borrador 
lenn=len(df_aleatorio2)
i=random.randint(1, lenn) 

for i in range (longitud, longitud2):
	df_aleatorio2.loc[i, 'Idioma']=df_aleatorio.loc[i, 'Idioma']
	df_aleatorio2.loc[i, 'Usuario']=df_aleatorio.loc[i, 'Usuario']
	df_aleatorio2.loc[i, 'Contrasena']=df_aleatorio.loc[i, 'Contrasena']
	df_aleatorio2.loc[i, 'Contenedor']=df_aleatorio.loc[i, 'Contenedor']
	df_aleatorio2.loc[i, 'Incoterms']=df_aleatorio.loc[i, 'Incoterms']
	df_aleatorio2.loc[i, 'IdCliente']=df_aleatorio.loc[i, 'IdCliente']
	df_aleatorio2.loc[i, 'IdDireccion']=df_aleatorio.loc[i, 'IdDireccion']
	df_aleatorio2.loc[i, 'IdPuerto']=0
	df_aleatorio2.loc[i, 'CapacidadDescarga']=0
	df_aleatorio2.loc[i, 'CapacidadDisponible']=0
	df_aleatorio2.loc[i, 'ProductoSoportaPeso']=0
	df_aleatorio2.loc[i, 'PesoProducto1']=0
	df_aleatorio2.loc[i, 'ProductoNoSoportaPeso']=0
	df_aleatorio2.loc[i, 'PesoProducto2']=0
	df_aleatorio2.loc[i, 'Iterar']=0
	df_aleatorio2.loc[i, 'Resultado']= "Prueba fallida"

#Crear dataframe con filas aleatorias
df_aleatorio3=pd.DataFrame()
nag=len(df_aleatorio2)
df_aleatorio3 = df_aleatorio2.sample(nag)

export_excel2 = df_aleatorio3.to_excel (r'Pedidos/DataExcel/Crear_Pedido40pies.xlsx', index = None, header=True)

#Borrado de archivos
os.remove('TestDataIdioma.xlsx')
os.remove('TestDataClient.xlsx')
os.remove('TestDataContenedor.xlsx') 
os.remove('TestDataIncoterms.xlsx')
os.remove('TestDataProductNo.xlsx')
os.remove('TestDataProductSop.xlsx')
os.remove('Testfusion.xlsx')
os.remove('TestDataUser.xlsx')

#Se valida y se crea el archivo que necesita Katalon para guardar los pedidos
if os.path.isfile("Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx"):
    print("Ya esta creado: Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx")
else:
	wb7 = openpyxl.Workbook() 
	wb7.save('Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx') 
	print("Archivo creado: Pedidos/DataExcel/Pedidos_Creados_Katalon.xlsx")