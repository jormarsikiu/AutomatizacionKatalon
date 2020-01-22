# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
import datetime
from openpyxl.utils import get_column_letter


idpedidos = pd.read_excel(r"Pedidos_Creados_Katalon.xlsx")

df_pedidos= pd.DataFrame(idpedidos, columns= ['IdPedido'])
#n=len(df_pedidos)

l=df_pedidos.iloc[:, 0].tolist()

#Conexion con la base de fatos
try:
	server = '35.196.201.168' 
	database = 'PETROCHEMICAL_PRUEBAS' 
	username = 'KAIZEN' 
	password = 'SYSERP2016-9#' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto")
except Exception as e:
	print ("No Conecto")
	raise e

data = pd.read_sql('select P.ID AS IdPedido, P.ID_CLIENTE AS Id_Cliente, P.FECHA_CREACION AS Creado_desde, E.NOMBRE As Estatus_Pedido from PEDIDO_MOVIL AS P, ESTATUS_PEDIDO AS E WHERE P.ID_ESTATUS = E.ID AND P.ID in (' + ','.join((str(n) for n in l)) + ')',cnxn)

data.to_excel('Ventas.xlsx')
ventas = pd.read_excel(r"Ventas.xlsx")

df_ventas= pd.DataFrame(ventas, columns= ['IdPedido', 'Id_Cliente', 'Estatus_Pedido', 'Creado_desde'])
df_ventas['Creado_desde'] = df_ventas['Creado_desde'].dt.strftime('%d/%m/%Y')
df_ventas['Creado_hasta'] = df_ventas['Creado_desde']
export_excel4 = df_ventas.to_excel (r'Ventas.xlsx', index = None, header=True)

wb = openpyxl.load_workbook('Ventas.xlsx')
sheet = wb.worksheets[0]
n=len(sheet['A'])


#################################################################################
#Escribir el Usuario y contrasena
wb5 = openpyxl.Workbook() 
sheet5 = wb5.active
sheet5.cell(row=1, column=1).value ="Usuario"
sheet5.cell(row=1, column=2).value ="Contrasena"
count=0

for rows in sheet5.iter_cols(min_col=1, max_col=1, min_row=2, max_row=n):
	for row in rows:
		row.value="automated.test@grupokaizen.net"

for rows in sheet5.iter_cols(min_col=2, max_col=2, min_row=2, max_row=n):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb5.save("TestDataUser.xlsx") 

usuarios = pd.read_excel(r"TestDataUser.xlsx")
df_usuarios =  pd.DataFrame(usuarios, columns= ['Usuario', 'Contrasena'])
#################################################################################

#Escribir el Estado del pedido 
wb1 = openpyxl.Workbook() 
sheet1 = wb1.active
sheet1.cell(row=1, column=1).value ="Estado_Pedido"

l=(len(df_ventas))+1
count=0
for rows1 in sheet1.iter_cols(min_col=1, max_col=1, min_row=2, max_row=l):
	for row1 in rows1:
			count = count+1
			if count%2 == 0:
				row1.value="Aprobar"
			else:
				row1.value="Denegar"
wb1.save("TestEstadoPedido.xlsx") 

#Concatenar los dataframe en uno 
estado = pd.read_excel(r"TestEstadoPedido.xlsx")
df_estado= pd.DataFrame(estado, columns= ['Estado_Pedido'])
df_fusion2=pd.concat([df_usuarios, df_ventas, df_estado], axis=1)
export_excel2 = df_fusion2.to_excel (r'AprobacionV.xlsx', index = None, header=True)

#Filtrado de los pedidos Creados quitando los borradores 
df_fusion3= pd.DataFrame()
for i in range (0, len(df_fusion2)):
	if (df_fusion2.loc[i, 'Estatus_Pedido']=="Creado"):
		df_fusion3.loc[i, 'Usuario']=df_fusion2.loc[i, 'Usuario']
		df_fusion3.loc[i, 'Contrasena']=df_fusion2.loc[i, 'Contrasena']
		df_fusion3.loc[i, 'IdPedido']=df_fusion2.loc[i, 'IdPedido']
		df_fusion3.loc[i, 'Id_Cliente']=df_fusion2.loc[i, 'Id_Cliente']
		df_fusion3.loc[i, 'Estatus_Pedido']=df_fusion2.loc[i, 'Estatus_Pedido']
		df_fusion3.loc[i, 'Creado_desde']=df_fusion2.loc[i, 'Creado_desde']
		df_fusion3.loc[i, 'Creado_hasta']=df_fusion2.loc[i, 'Creado_hasta']
		df_fusion3.loc[i, 'Estado_Pedido']=df_fusion2.loc[i, 'Estado_Pedido']

export_excel3 = df_fusion3.to_excel (r'AprobacionVentas.xlsx', index = None, header=True)

os.remove('TestEstadoPedido.xlsx')
os.remove('ventas.xlsx')
os.remove('AprobacionV.xlsx')