# -*- coding: utf-8 -*-
#!/usr/bin/python
import pyodbc
import openpyxl
import pandas as pd
import os
import random

#Conexion con la base de fatos
try:
	server = '35.196.201.168' 
	database = 'PETROCHEMICAL_PRUEBAS' 
	username = 'KAIZEN' 
	password = 'SYSERP2016-9#' 
	cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
	print ("Conecto a la BD")
except Exception as e:
	print ("No Conecto a la BD")
	raise e

data = pd.read_sql('SELECT ID as IdProducto, CODIGO as CodProducto, NOMBRE as Producto, EXISTENCIA as Existencia, ID_GRUPO as IdGrupo, EN_VENTA as EnVenta FROM PRODUCTO WHERE ELIMINADO = 0 AND ACTIVO = 1 AND ID_GRUPO IN (6, 7, 32) AND EXISTENCIA>0',cnxn)
data2= pd.read_sql('SELECT ID as IdProducto, CODIGO as CodProducto, NOMBRE as Producto, EXISTENCIA as Existencia, ID_GRUPO as IdGrupo, EN_VENTA as EnVenta FROM PRODUCTO WHERE ELIMINADO = 0 AND ACTIVO = 1 AND ID_GRUPO IN (6, 7, 32) AND EXISTENCIA>0 AND EN_VENTA=1',cnxn)


data.to_excel('Producto.xlsx')
data2.to_excel('ProductoEnVenta.xlsx')

producto1=pd.read_excel(r"Producto.xlsx")
df_producto1=  pd.DataFrame(producto1, columns= ['IdProducto', 'CodProducto', 'Producto', 'Existencia', 'IdGrupo', 'EnVenta'])

productoEnVenta=pd.read_excel(r"ProductoEnVenta.xlsx")
df_productoEnVenta=pd.DataFrame(productoEnVenta, columns= ['IdProducto', 'CodProducto', 'Producto',  'Existencia', 'IdGrupo', 'EnVenta',])

nc=len(df_producto1)
df_producto1 = df_producto1.sample(nc)
df_producto1 = df_producto1.reset_index(drop=True)
df_producto1 = df_producto1.iloc[0:10]

na=len(df_productoEnVenta)
df_productoEnVenta = df_productoEnVenta.sample(na)
df_productoEnVenta = df_productoEnVenta.reset_index(drop=True)
df_productoEnVenta = df_productoEnVenta.iloc[0:2]

df_producto=pd.concat([df_productoEnVenta, df_producto1], ignore_index=True)


for i in range (0, len(df_producto)):
	if (df_producto.loc[i, 'EnVenta']==True):
		df_producto.loc[i, 'EnVenta']=1
	elif (df_producto.loc[i, 'EnVenta']==False):
		df_producto.loc[i, 'EnVenta']=0
		
#Aleatorios de cantidad minima y cantidad maxima
for i in range (0, len(df_producto)):
	if (df_producto.loc[i, 'IdGrupo']==6):
		df_producto.loc[i, 'CantidadMinima'] = random.randint(10, 100)
	elif (df_producto.loc[i, 'IdGrupo']==7):
		df_producto.loc[i, 'CantidadMinima'] = random.randint(0, 20)
	elif (df_producto.loc[i, 'IdGrupo']==32):
		df_producto.loc[i, 'CantidadMinima'] = random.randint(0, 4000)

for i in range (0, len(df_producto)):
	if (df_producto.loc[i, 'IdGrupo']==6):
		df_producto.loc[i, 'CantidadMaxima'] = random.randint(100, 100000)
	elif (df_producto.loc[i, 'IdGrupo']==7):
		df_producto.loc[i, 'CantidadMaxima'] = random.randint(20, 250)
	elif (df_producto.loc[i, 'IdGrupo']==32):
		df_producto.loc[i, 'CantidadMaxima'] = random.randint(4000, 20000)

#Asignar el almacen de materia prima o de producto terminado
AlmacenMP = [1, 6]
AlmacenPT = [2, 7]

for i in range (0, len(df_producto)):
	if ((df_producto.loc[i, 'IdGrupo']==6)  or (df_producto.loc[i, 'IdGrupo']==32)):
		df_producto.loc[i, 'AgregarAlmacen1'] = random.choice(AlmacenMP)
	elif (df_producto.loc[i, 'IdGrupo']==7):
		df_producto.loc[i, 'AgregarAlmacen1'] = random.choice(AlmacenPT)

for i in range (0, len(df_producto)):
	if ((df_producto.loc[i, 'IdGrupo']==6)  or (df_producto.loc[i, 'IdGrupo']==32)):
		if (df_producto.loc[i, 'AgregarAlmacen1']==AlmacenMP[0]):
			df_producto.loc[i, 'AgregarAlmacen2'] = AlmacenMP[1]
		elif (df_producto.loc[i, 'AgregarAlmacen1']==AlmacenMP[1]):
			df_producto.loc[i, 'AgregarAlmacen2'] = AlmacenMP[0]
	elif (df_producto.loc[i, 'IdGrupo']==7):
		if (df_producto.loc[i, 'AgregarAlmacen1']==AlmacenPT[0]):
			df_producto.loc[i, 'AgregarAlmacen2'] = AlmacenPT[1]
		elif (df_producto.loc[i, 'AgregarAlmacen1']==AlmacenPT[1]):
			df_producto.loc[i, 'AgregarAlmacen2'] =AlmacenPT[0]

#Cambiar el nombre del grupo
for i in range (0, len(df_producto)):
	if (df_producto.loc[i, 'IdGrupo']==6):
		df_producto.loc[i, 'IdGrupo'] = 'Materia Prima'
	elif (df_producto.loc[i, 'IdGrupo']==7):
		df_producto.loc[i, 'IdGrupo'] = 'Producto Terminado'
	elif (df_producto.loc[i, 'IdGrupo']==32):
		df_producto.loc[i, 'IdGrupo'] = 'MP: Liquidos'

#for i in range (0, len(df_producto)):
#	df_producto.loc[i, 'CantidadAnadir'] = random.randint(0, 100)


export_excel = df_producto.to_excel (r'Producto.xlsx', index = None, header=True)

#################################################################################

#Escribir el Usuario y contrasena
wb = openpyxl.Workbook() 
sheet = wb.active
sheet.cell(row=1, column=1).value ="Idioma"
sheet.cell(row=1, column=2).value ="Usuario"
sheet.cell(row=1, column=3).value ="Contrasena"

count=0
count5=0

lg=(len(df_producto))+1
for rows in sheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=lg):
	for row in rows:
			count = count+1
			if count%2 == 0:
				row.value="1"
			else:
				row.value="1"	

for rows in sheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=lg):
	for row in rows:
		row.value="automated.test@grupokaizen.net"

for rows in sheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=lg):
	for row in rows:
		row.value="aeHFOx8jV/A="

wb.save("TestDataUser.xlsx") 

usuarios = pd.read_excel(r"TestDataUser.xlsx")
df_usuarios =  pd.DataFrame(usuarios, columns= ['Idioma', 'Usuario', 'Contrasena'])


#################################################################################

df_almacen=pd.concat([df_usuarios, df_producto], axis=1)

export_excel = df_almacen.to_excel (r'Pedidos/DataExcel/Inventario/ProductoAlmacen.xlsx', index = None, header=True)

os.remove('Producto.xlsx')
os.remove('TestDataUser.xlsx')
os.remove('ProductoEnVenta.xlsx')





